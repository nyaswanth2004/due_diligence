import csv
import io

from app.ingestion.extractors.base import BaseExtractor, ExtractionResult, PageContent


class SpreadsheetExtractor(BaseExtractor):
    """Extracts row content from spreadsheet files.

    Each worksheet is treated as a "page" so chunk provenance maps to a sheet.
    Supports .xlsx (openpyxl) and .csv (stdlib). Legacy .xls requires the
    optional `xlrd` package and is handled when available.
    """

    def extract(self, content: bytes, filename: str) -> ExtractionResult:
        lower = filename.lower()
        if lower.endswith(".csv"):
            return self._extract_csv(content)
        if lower.endswith(".xlsx"):
            return self._extract_xlsx(content)
        if lower.endswith(".xls"):
            return self._extract_xls(content)
        raise ValueError(f"unsupported spreadsheet type: {filename}")

    def _extract_xlsx(self, content: bytes) -> ExtractionResult:
        from openpyxl import load_workbook  # noqa: PLC0415

        pages: list[PageContent] = []
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        for sheet_index, name in enumerate(workbook.sheetnames, start=1):
            sheet = workbook[name]
            lines = [f"[SHEET] {name}"]
            for row in sheet.iter_rows(values_only=True):
                cleaned = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if cleaned:
                    lines.append(" | ".join(cleaned))
            pages.append(PageContent(page_number=sheet_index, text="\n".join(lines).strip()))
        return ExtractionResult(pages=[p for p in pages if p.text])

    def _extract_csv(self, content: bytes) -> ExtractionResult:
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        lines = []
        for row in reader:
            cleaned = [c.strip() for c in row if c.strip()]
            if cleaned:
                lines.append(" | ".join(cleaned))
        body = "\n".join(lines).strip()
        pages = [PageContent(page_number=1, text=body)] if body else []
        return ExtractionResult(pages=pages)

    def _extract_xls(self, content: bytes) -> ExtractionResult:
        try:
            import xlrd  # noqa: PLC0415
        except ImportError as exc:
            raise ValueError(
                "Legacy .xls files require `pip install xlrd`; convert to .xlsx or .csv instead."
            ) from exc
        workbook = xlrd.open_workbook(file_contents=content)
        pages: list[PageContent] = []
        for sheet_index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_index)
            lines = [f"[SHEET] {sheet.name}"]
            for row_index in range(sheet.nrows):
                cleaned = [
                    str(sheet.cell_value(row_index, col)).strip()
                    for col in range(sheet.ncols)
                    if str(sheet.cell_value(row_index, col)).strip()
                ]
                if cleaned:
                    lines.append(" | ".join(cleaned))
            pages.append(PageContent(page_number=sheet_index + 1, text="\n".join(lines).strip()))
        return ExtractionResult(pages=[p for p in pages if p.text])
